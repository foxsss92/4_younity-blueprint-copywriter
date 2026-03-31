import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ============================================================
# ROMANIAN TRANSLATIONS — Retargeting / Follow-up Headlines
# Translated in YOUnity brand voice:
# empowering, scientifically grounded, spiritually resonant
# ============================================================

headlines = [
    {
        "english": "Go Deeper Into The Human Blueprint",
        "romanian": "Pătrunzi Mai Adânc În Blueprintul Uman",
    },
    {
        "english": "Make The Shift Real This Time",
        "romanian": "Fă Schimbarea Reală De Data Aceasta",
    },
    {
        "english": "Come Back To Your Center",
        "romanian": "Revino La Centrul Tău",
    },
    {
        "english": "Go Beyond The Free Training",
        "romanian": "Mergi Dincolo De Seria Gratuită",
    },
    {
        "english": "Turn Insight Into Inner Change",
        "romanian": "Transformă Revelația În Schimbare Interioară",
    },
]

# ============================================================
# LOAD WORKBOOK & APPEND NEW SECTION
# ============================================================

wb_path = "/home/user/4_younity-blueprint-copywriter/Blueprint/brg-cc-k3/brg-cc-k3 Facebook Ad Copy.xlsx"
wb = openpyxl.load_workbook(wb_path)
ws = wb.active

align_top_wrap = Alignment(vertical="top", wrap_text=True)
font_bold = Font(bold=True)

def make_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

fill_adset    = make_fill("BDD7EE")
fill_varlabel = make_fill("E2EFDA")

current_row = ws.max_row + 2  # leave a gap after existing content

# --- Section header ---
ws.cell(row=current_row, column=2, value="RETARGETING / FOLLOW-UP HEADLINES — Romanian Translations").font = font_bold
ws.cell(row=current_row, column=2).fill = fill_adset
ws.cell(row=current_row, column=2).alignment = align_top_wrap
ws.cell(row=current_row, column=3, value="Original (English)").font = font_bold
ws.cell(row=current_row, column=3).fill = fill_adset
ws.cell(row=current_row, column=3).alignment = align_top_wrap
ws.cell(row=current_row, column=4, value="Romanian").font = font_bold
ws.cell(row=current_row, column=4).fill = fill_adset
ws.cell(row=current_row, column=4).alignment = align_top_wrap
ws.row_dimensions[current_row].height = 20
current_row += 1

# --- Each headline ---
for h in headlines:
    ws.cell(row=current_row, column=3, value=h["english"]).alignment = align_top_wrap
    ws.cell(row=current_row, column=4, value=h["romanian"]).alignment = align_top_wrap
    ws.row_dimensions[current_row].height = 22
    current_row += 1

wb.save(wb_path)
print("Saved translations to:", wb_path)
print()
print("Headlines translated (EN → RO):")
for h in headlines:
    print(f"  EN: {h['english']}")
    print(f"  RO: {h['romanian']}")
    print()
