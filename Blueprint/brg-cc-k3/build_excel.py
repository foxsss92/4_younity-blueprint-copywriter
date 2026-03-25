import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ============================================================
# ALL COPY CONTENT
# ============================================================

# ---- AD SET 1 — Problem Aware — AI is Replacing Your Humanity ----

ad_set_1_label = "AD SET 1 - Problem Aware - AI is Replacing Your Humanity"

meta_desc_1 = (
    "Your brain outperforms any machine ever built. Gregg Braden's free series reveals "
    "the biological truth AI cannot replicate."
)

# Variation 1a — conversational/empathetic, flowing paragraphs
hl1_1a = "You Are Not Becoming Obsolete"
hl2_1a = "The Most Advanced Technology Is You"
text_1a = (
    "There's a quiet dread spreading beneath the surface of everyday life. You don't always name it, "
    "but you feel it — that creeping sense that AI is moving into spaces once reserved for human minds, "
    "human judgment, human worth. Like maybe the things that made you valuable are slowly being written out.\n\n"
    "But here's what the headlines aren't telling you.\n\n"
    "Your brain processes 100 billion instructions per second. The most advanced microprocessors manage six billion. "
    "Your 50 trillion cells self-heal, self-regulate, and self-replicate without a single conscious instruction from you. "
    "Seven biological triggers — thought, feeling, breath, focus, movement, nutrition, emotion — "
    "reshape your biology in real time. You are not a user of technology. You are soft technology.\n\n"
    "Gregg Braden — scientist, five-time New York Times bestselling author — is hosting a free 3-day online series "
    "called The Human Blueprint. Join him to rediscover what you actually are.\n\n"
    "Register free. The series starts soon."
)
graphics_1a = (
    "Graphic Title: You Are the Most Advanced Technology on Earth\n\n"
    "Graphic Pre-header: AI processes six billion instructions per second. Your brain processes one hundred billion. "
    "The race was never close.\n\n"
    "Graphic CTA: Join the Free Series"
)

# Variation 1b — short punchy lines, one sentence per line, bold/direct
hl1_1b = "AI Can't Replace What You've Forgotten"
hl2_1b = "You Were Built for More Than This"
text_1b = (
    "Every week there's a new headline.\n"
    "Another job category gone.\n"
    "Another skill made redundant.\n"
    "Another reason to wonder where humans fit.\n\n"
    "But nobody's talking about what you actually are.\n\n"
    "Your brain runs 100 billion instructions per second.\n"
    "The best chip ever built manages six billion.\n"
    "Your cells self-heal without software updates.\n"
    "Your body is 50 trillion living electrical circuits.\n"
    "You change your own biology with a single shift in thought.\n\n"
    "You were not falling behind technology.\n"
    "You forgot what you already were.\n\n"
    "Gregg Braden is hosting a free 3-day series — The Human Blueprint — "
    "to give you back the truth that never left.\n\n"
    "Claim your free spot now."
)
graphics_1b = (
    "Graphic Title: The Machine Was Never the Threat\n\n"
    "Graphic Pre-header: You were built before AI. You process more, heal more, adapt more. "
    "The question was never competition — it was memory.\n\n"
    "Graphic CTA: Register Free Today"
)

# Variation 1c — storytelling/narrative
hl1_1c = "The Day We Started Building Machines to Remember"
hl2_1c = "What Ancient Peoples Knew That We Forgot"
text_1c = (
    "Years ago, Gregg Braden stood with a native elder at the edge of the ancient ruins of Chaco Canyon. "
    "Looking out over the landscape, the elder said something that has stayed with him ever since.\n\n"
    "He said that long ago, humans lost their connection to the deep intelligence within themselves — "
    "their capacity to feel across distances, to know things beyond ordinary sight, to heal from the inside. "
    "And then, he said, they began building machines to recreate what they had lost.\n\n"
    "The internet — a mirror of our lost capacity for collective consciousness.\n"
    "AI — a reflection of the intelligence we stopped trusting in ourselves.\n"
    "Virtual reality — a pale copy of the inner worlds we once had direct access to.\n\n"
    "We didn't create technology to evolve. We created it because we forgot who we were.\n\n"
    "Gregg Braden's free 3-day series, The Human Blueprint, is an invitation to remember. "
    "The science is real. The timing is urgent. The series is free.\n\n"
    "Save your seat now."
)
graphics_1c = (
    "Graphic Title: We Built the Machines Because We Forgot\n\n"
    "Graphic Pre-header: A native elder at Chaco Canyon told Gregg Braden the truth about why technology exists — "
    "and what it says about who we really are.\n\n"
    "Graphic CTA: Join the Free Series"
)

# ---- AD SET 2 — Problem Aware — Feeling Disconnected Despite Seeking ----

ad_set_2_label = "AD SET 2 - Problem Aware - Feeling Disconnected Despite Seeking"

meta_desc_2 = (
    "Understanding isn't enough — your heart has its own neural network. "
    "Gregg Braden's free series shows where real change actually takes root."
)

# Variation 2a — question-led
hl1_2a = "Why Knowing Isn't Enough"
hl2_2a = "The Gap Between Insight and Living It"
text_2a = (
    "How many times have you understood something completely — and still felt it slip away the moment life got hard?\n\n"
    "You know the pattern. You know the trigger. You've read the books, done the work. "
    "And then a conversation goes sideways, or the anxiety rises in the middle of the night, "
    "and the version of yourself you've been building seems suddenly far away.\n\n"
    "What if it isn't failure? What if it's anatomy?\n\n"
    "Your heart has 40,000 neurons and its own independent neural network. "
    "It processes emotional reality separately from your brain — and it sends more signals up to the brain "
    "than the brain sends down to it. When heart and brain are not in coherence, "
    "no insight, no intention, no amount of understanding can fully bridge the gap.\n\n"
    "You don't need more information. You need the signal to lock in.\n\n"
    "Gregg Braden's free 3-day series, The Human Blueprint, shows you exactly where change actually takes root. "
    "Register free and join him live."
)
graphics_2a = (
    "Graphic Title: Your Heart Knows What Your Mind Can't Fix\n\n"
    "Graphic Pre-header: The heart has 40,000 neurons and its own neural network — "
    "and it's been sending signals your brain hasn't been receiving.\n\n"
    "Graphic CTA: Join the Free Series"
)

# Variation 2b — flowing paragraphs, reflective
hl1_2b = "Something Keeps Pulling You Back"
hl2_2b = "The Pattern That Lives Below the Surface"
text_2b = (
    "You make a shift — a real one. You feel it. Something loosens. You move differently through your days.\n\n"
    "And then, gradually, without any single dramatic moment, you find yourself back. "
    "Same reactions. Same internal weather. Same old version of yourself that you thought you'd grown past.\n\n"
    "This isn't weakness. It isn't lack of commitment. It's biology that hasn't caught up to belief.\n\n"
    "The body holds patterns the way it holds memories — encoded at a cellular level, below the reach "
    "of conscious thought and intention. You can understand your way to the edge of change. "
    "But understanding alone doesn't rewrite the body's default state.\n\n"
    "Gregg Braden has spent three decades at the intersection of science and ancient wisdom, "
    "mapping exactly where change takes root — not just in the mind, but in the biology. "
    "His free 3-day series, The Human Blueprint, shows you the mechanism, not just the message.\n\n"
    "Register free. Come find where it actually shifts."
)
graphics_2b = (
    "Graphic Title: The Pattern Is Not a Flaw. It's a Signal.\n\n"
    "Graphic Pre-header: Real change doesn't start in the mind. Gregg Braden shows where it actually begins — "
    "and why your body keeps defaulting back.\n\n"
    "Graphic CTA: Register Free"
)

# Variation 2c — short punchy lines, bold
hl1_2c = "You're Not Broken. You're Incomplete."
hl2_2c = "One Missing Piece Changes Everything"
text_2c = (
    "You've done the inner work.\n"
    "You've read the books.\n"
    "You understand the patterns.\n\n"
    "And you still default.\n\n"
    "Here's what nobody told you:\n\n"
    "Your heart has 40,000 neurons.\n"
    "It processes reality before your brain does.\n"
    "It sends more signals up than the brain sends down.\n"
    "When heart and brain are out of sync, your immune response drops.\n"
    "Stress hormones stay elevated.\n"
    "And no mindset work in the world closes that gap.\n\n"
    "This isn't a motivation problem.\n"
    "It's a coherence problem.\n\n"
    "Gregg Braden's free 3-day series — The Human Blueprint — "
    "gives you the missing piece the self-help industry never mentioned.\n\n"
    "Claim your free spot. It starts soon."
)
graphics_2c = (
    "Graphic Title: One Missing Piece Changes Everything\n\n"
    "Graphic Pre-header: The heart has its own nervous system — "
    "and when it's out of sync with your brain, no amount of inner work sticks.\n\n"
    "Graphic CTA: Join Free Now"
)

# ---- AD SET 3 — Solution Aware — Heart-Brain Coherence ----

ad_set_3_label = "AD SET 3 - Solution Aware - Heart-Brain Coherence"

meta_desc_3 = (
    "Heart-brain coherence triggers 1,300+ biochemical reactions, drops cortisol by 23%, "
    "and doubles DHEA. Gregg Braden's free series shows you how to sustain it."
)

# Variation 3a — flowing paragraphs
hl1_3a = "Heart-Brain Coherence Is Only the Beginning"
hl2_3a = "What Happens When the Signal Locks In"
text_3a = (
    "You've felt it before — that rare, crystalline moment when something inside you settles. "
    "A clarity that doesn't feel mental. A steadiness that comes from somewhere deeper than thought. "
    "And then, almost without noticing, it was gone.\n\n"
    "That was coherence. And it's far more than a feeling.\n\n"
    "When the heart and brain enter sustained coherence, the body initiates over 1,300 biochemical reactions. "
    "Immune antibodies — secretory IgA — surge. Cortisol drops by 23%. DHEA, the body's master resilience hormone, "
    "increases by 100%. Deep intuition sharpens. Memory consolidates. "
    "The electromagnetic field of the heart — 5,000 times stronger than the brain's — stabilizes.\n\n"
    "This is not a relaxation technique. This is a different biological frequency entirely.\n\n"
    "The challenge is sustaining it — moving from those fleeting moments into a lived, embodied state. "
    "That's exactly what Gregg Braden teaches in his free 3-day online series, The Human Blueprint.\n\n"
    "Register free and learn what changes when the signal stays."
)
graphics_3a = (
    "Graphic Title: 1,300 Biochemical Reactions. One Coherent Signal.\n\n"
    "Graphic Pre-header: Sustained heart-brain coherence doesn't just feel good — "
    "it restructures your immune system, stress response, and intuition from the inside out.\n\n"
    "Graphic CTA: Join the Free Series"
)

# Variation 3b — short punchy lines
hl1_3b = "You've Felt It. Now Learn to Keep It."
hl2_3b = "The Science of Sustained Coherence"
text_3b = (
    "That moment of sudden quiet clarity.\n"
    "When everything slows down and you just... know.\n"
    "You've felt it.\n\n"
    "Science has a name for it: heart-brain coherence.\n\n"
    "Your heart's electromagnetic field extends three feet outside your body.\n"
    "It has four distinct neural pathways to the brain.\n"
    "In coherence, it triggers 1,300+ biochemical shifts.\n"
    "Cortisol drops 23%.\n"
    "DHEA — your resilience hormone — doubles.\n"
    "Immune function rises measurably.\n\n"
    "The technique takes three minutes.\n"
    "The effects last far longer.\n\n"
    "Gregg Braden has spent decades mapping this — and in his free 3-day series, "
    "The Human Blueprint, he shows you the full picture: what coherence really is, "
    "why it matters now more than ever, and how to stop losing it.\n\n"
    "Register free. The series is live soon."
)
graphics_3b = (
    "Graphic Title: You've Felt It. Now Learn to Keep It.\n\n"
    "Graphic Pre-header: Three minutes. Four neural pathways. Over 1,300 biochemical reactions. "
    "Heart-brain coherence is the most powerful state you haven't been taught to sustain.\n\n"
    "Graphic CTA: Register Free"
)

# Variation 3c — storytelling, ancient wisdom meets science
hl1_3c = "Ancient Traditions Said It First. Science Just Caught Up."
hl2_3c = "The Intelligence Living in Your Chest"
text_3c = (
    "For thousands of years, ancient traditions — from the Vedic schools of India to the mystery schools "
    "of Egypt, from indigenous wisdom keepers to the early Christian mystics — placed the heart, "
    "not the brain, at the center of human intelligence and spiritual knowing.\n\n"
    "Modern science dismissed this as poetic metaphor. Until it couldn't.\n\n"
    "The heart contains 40,000 neurons — enough to constitute an independent nervous system. "
    "Its electromagnetic field is 5,000 times stronger than the brain's. "
    "It sends more neural signals upward to the brain than the brain sends down to it. "
    "It perceives, remembers, and processes emotional reality on its own terms.\n\n"
    "And when it enters coherence with the brain, more than 1,300 biochemical reactions activate — "
    "rebalancing hormones, lifting immune function, and unlocking a quality of intuition that has no "
    "equivalent in the cognitive mind.\n\n"
    "The ancients weren't being metaphorical. They were being precise.\n\n"
    "Gregg Braden bridges this knowing with modern science in his free 3-day series, The Human Blueprint. "
    "Register free and discover what your heart has always known."
)
graphics_3c = (
    "Graphic Title: The Heart Knew Before Science Did\n\n"
    "Graphic Pre-header: Ancient traditions placed the heart at the center of intelligence for thousands of years. "
    "Modern science just confirmed why — 40,000 neurons, 5,000x stronger electromagnetic field, 1,300 biochemical reactions.\n\n"
    "Graphic CTA: Join the Free Series"
)

# ---- AD SET 4 — Solution Aware — DNA Carries a Divine Message ----

ad_set_4_label = "AD SET 4 - Solution Aware - DNA Carries a Divine Message"

meta_desc_4 = (
    "Scientists mapped DNA bases to atomic weights — and found ancient Hebrew letters spelling "
    "\"God eternal within the body\" in every one of your 50 trillion cells."
)

# Variation 4a — reflective/introspective, flowing paragraphs
hl1_4a = "The Message Hidden in Every Cell of Your Body"
hl2_4a = "Your DNA Was Written With Intention"
text_4a = (
    "What if the most sacred text ever written isn't in a library somewhere — "
    "but inside every cell of your body?\n\n"
    "Scientists mapping the chemical bases of human DNA to their atomic weights discovered "
    "a precise correspondence: each base maps to an ancient Hebrew letter. "
    "Those letters, read in sequence, spell a phrase — \"God eternal within the body\" — "
    "encoded into the nucleotide structure of every human cell. "
    "The probability of this occurring by chance: 0.00042%.\n\n"
    "Then there's Chromosome 2 — a fusion of two ancestral chromosomes that cannot have occurred through "
    "natural random processes. Embedded within it: the TBR1 gene, linked scientifically to empathy, "
    "intuition, and compassion. The very qualities that define what it means to be most deeply human.\n\n"
    "You are not the product of random processes. You are not ordinary. "
    "Every one of your 50 trillion cells carries the same encoded declaration.\n\n"
    "Gregg Braden unpacks this discovery in his free 3-day series, The Human Blueprint. "
    "Register free and read the message written into you."
)
graphics_4a = (
    "Graphic Title: The Message Is Inside Every Cell\n\n"
    "Graphic Pre-header: Scientists mapped human DNA bases to atomic weights and found ancient Hebrew letters "
    "spelling \"God eternal within the body\" — in every one of your 50 trillion cells.\n\n"
    "Graphic CTA: Join the Free Series"
)

# Variation 4b — conversational/empathetic
hl1_4b = "You Were Not an Accident"
hl2_4b = "The Intentional Design Inside Your Biology"
text_4b = (
    "There are moments — usually quiet ones — when a thought surfaces that most people never say out loud. "
    "A feeling of smallness. A sense that in the grand sweep of everything, you don't particularly matter. "
    "That your presence here is random, unremarkable, coincidental.\n\n"
    "What if the science itself argues otherwise?\n\n"
    "Chromosome 2 in the human genome is a fusion of two ancestral chromosomes — "
    "a fusion that cannot happen through ordinary random mutation. It was, by all scientific accounts, "
    "a deliberate structural event. Embedded within it: the TBR1 gene, the biological root of empathy, "
    "intuition, and compassion. The things that make you most irreducibly yourself.\n\n"
    "And in the chemical bases of your DNA — when mapped to their atomic weights — "
    "ancient Hebrew letters emerge, reading: \"God eternal within the body.\" "
    "In every one of your 50 trillion cells. Every one. Every moment.\n\n"
    "You were not an accident. The message is inside you — written long before you were born.\n\n"
    "Gregg Braden explores this in his free series, The Human Blueprint. Register and come find yourself."
)
graphics_4b = (
    "Graphic Title: You Were Not an Accident\n\n"
    "Graphic Pre-header: Your Chromosome 2 carries a fusion that can't happen randomly, "
    "and your DNA encodes a message that changes everything about what you think you are.\n\n"
    "Graphic CTA: Register Free"
)

# Variation 4c — bold/punchy lines
hl1_4c = "Science Just Confirmed What Mystics Always Said"
hl2_4c = "50 Trillion Encoded Messages"
text_4c = (
    "2007. Scientists confirm DNA stores and transmits information.\n\n"
    "They map DNA's chemical bases to their atomic weights.\n"
    "The atomic weights correspond precisely to ancient Hebrew letters.\n"
    "The letters spell a phrase.\n"
    "That phrase: \"God eternal within the body.\"\n\n"
    "In every human cell.\n"
    "50 trillion times over in your body alone.\n"
    "Probability of coincidence: 0.00042%.\n\n"
    "Then there's Chromosome 2.\n"
    "A fusion of two ancestral chromosomes.\n"
    "Scientifically impossible through random mutation.\n"
    "Containing the TBR1 gene — the biological seat of empathy, intuition, compassion.\n\n"
    "Gregg Braden — earth scientist, author of The Divine Matrix — "
    "calls this the most significant discovery of our time.\n\n"
    "He's presenting it free, for three days, in The Human Blueprint.\n\n"
    "Claim your seat. The series starts soon."
)
graphics_4c = (
    "Graphic Title: 50 Trillion Encoded Messages\n\n"
    "Graphic Pre-header: DNA bases mapped to atomic weights. Ancient Hebrew letters. "
    "\"God eternal within the body\" — written into every cell. Probability of coincidence: 0.00042%.\n\n"
    "Graphic CTA: Join Free Now"
)

# ---- AD SET 5 — Most Aware — Gregg Braden's Free Series ----

ad_set_5_label = "AD SET 5 - Most Aware - Gregg Braden's Free Series"

meta_desc_5 = (
    "Three days. Gregg Braden live and free. The Human Blueprint series covers DNA, "
    "heart-brain coherence, and navigating AI without losing what makes you human."
)

# Variation 5a — flowing paragraphs, for existing followers
hl1_5a = "Gregg Braden's Free 3-Day Series Is Open"
hl2_5a = "The Human Blueprint — Live and Free"
text_5a = (
    "If Gregg Braden's work has stayed with you — if the connections he draws between ancient knowing "
    "and modern science have shifted the way you understand yourself — then this is the next conversation.\n\n"
    "The Human Blueprint is a free 3-day online series, and it brings together the threads "
    "he's been weaving for three decades into one integrated framework.\n\n"
    "Day one: why your body is the most advanced technology that has ever existed — "
    "and what that means in a world accelerating toward human-AI convergence.\n"
    "Day two: the encoded message inside your DNA, and what it reveals about your origin, "
    "your design, and your deepest capacities.\n"
    "Day three: heart-brain coherence as a biological superpower — the science, the technique, "
    "and the capabilities no machine will ever replicate.\n\n"
    "This is not a replay of what you already know. This is where it goes deeper.\n\n"
    "Register free. The series is live soon and seats are limited."
)
graphics_5a = (
    "Graphic Title: The Human Blueprint — Free 3-Day Series with Gregg Braden\n\n"
    "Graphic Pre-header: Three days. Three sessions. The science of who you really are — "
    "brought together in one free online series.\n\n"
    "Graphic CTA: Save Your Free Seat"
)

# Variation 5b — short punchy lines
hl1_5b = "Three Days. Gregg Braden. Free."
hl2_5b = "The Human Blueprint Series"
text_5b = (
    "You know the work.\n"
    "You've read the books.\n"
    "You know what Gregg Braden does with science and ancient wisdom.\n\n"
    "This is three decades of research in one free series.\n\n"
    "Day 1: You are the most advanced technology on Earth.\n"
    "Your brain runs 100 billion instructions per second.\n"
    "AI manages six billion.\n"
    "Your cells self-heal without a software update.\n\n"
    "Day 2: Your DNA carries a message.\n"
    "Encoded in every cell.\n"
    "\"God eternal within the body.\"\n"
    "50 trillion times over.\n\n"
    "Day 3: Heart-brain coherence.\n"
    "1,300+ biochemical reactions.\n"
    "Deep intuition. Immune surge. Cortisol drop.\n"
    "Capabilities no machine will replicate.\n\n"
    "The Human Blueprint. Free. Live. Three days.\n\n"
    "Register now before it fills."
)
graphics_5b = (
    "Graphic Title: Three Days. Gregg Braden. Free.\n\n"
    "Graphic Pre-header: The Human Blueprint — a free 3-day online series covering your biology, "
    "your DNA, and your untapped capacities. Everything you've been waiting to go deeper on.\n\n"
    "Graphic CTA: Register Free Now"
)

# Variation 5c — reflective, for deep readers
hl1_5c = "If His Books Changed How You See Yourself…"
hl2_5c = "This Series Goes Even Further"
text_5c = (
    "There's a particular kind of shift that happens when you encounter ideas that don't just inform you — "
    "they relocate you. You read a page and realize you've been standing in the wrong room your entire life. "
    "If Gregg Braden's work has done that for you, you already know what we mean.\n\n"
    "The Human Blueprint is where understanding becomes embodied knowing.\n\n"
    "Over three live sessions, Braden doesn't just present the research — he takes you through it "
    "in a way that lands in the body, not just the mind. The science of your biological superiority. "
    "The encoded message in your DNA. The heart-brain coherence state that changes your "
    "immune system, your stress chemistry, your intuition — and how to sustain it.\n\n"
    "This is for those who've been waiting to go further. Not just to understand more, "
    "but to become more of what they already are.\n\n"
    "The series is free. It's live. And it's the deepest he's gone in an open format.\n\n"
    "Register free. Come ready."
)
graphics_5c = (
    "Graphic Title: For Those Ready to Go Deeper\n\n"
    "Graphic Pre-header: Gregg Braden's free 3-day series isn't an introduction — "
    "it's the next level. Three sessions. Live. For those who've been waiting.\n\n"
    "Graphic CTA: Save My Free Seat"
)


# ============================================================
# BUILD THE WORKBOOK
# ============================================================

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "brg-cc-k3 Facebook Text"

# --- Styles ---
def make_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

fill_header   = make_fill("CCCCCC")
fill_adset    = make_fill("BDD7EE")
fill_metadesc = make_fill("FFFFC0")
fill_varlabel = make_fill("E2EFDA")

font_bold   = Font(bold=True)
font_italic = Font(italic=True)
font_normal = Font()

align_top_wrap = Alignment(vertical="top", wrap_text=True)

# Column widths  (A=1, B=2, C=3, D=4, E=5, F=6, G=7, H=8, I=9, J=10)
col_widths = {1: 3, 2: 28, 3: 65, 4: 65, 5: 65, 6: 65, 7: 65, 8: 65, 9: 65, 10: 3}
for col_idx, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col_idx)].width = width

def set_cell(ws, row, col, value, font=None, fill=None, alignment=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:      cell.font = font
    if fill:      cell.fill = fill
    if alignment: cell.alignment = alignment
    return cell

def apply_row_defaults(ws, row, cols=range(1, 11)):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.alignment = align_top_wrap

# ---- Row 1: empty ----
ws.row_dimensions[1].height = 6

# ---- Row 2: headers ----
header_vals = {
    2: 'What (Graphic / Text)',
    3: 'Original language',
    4: 'Translated language',
    5: 'Original Language2',
    6: 'Translated Round 2',
    7: 'Original Language3',
    8: 'Translated Round 3',
    9: 'Graphics Copy',
}
for col, val in header_vals.items():
    set_cell(ws, 2, col, val, font=font_bold, fill=fill_header, alignment=align_top_wrap)
ws.row_dimensions[2].height = 20

# ============================================================
# DATA: angles
# ============================================================
# Each angle: (adset_label, meta_desc, variations)
# Each variation: (var_id, hl1, hl2, text, graphics)

angles = [
    (
        ad_set_1_label, meta_desc_1,
        [
            ("1a", "1.1a", hl1_1a, hl2_1a, text_1a, graphics_1a),
            ("1b", "1.1b", hl1_1b, hl2_1b, text_1b, graphics_1b),
            ("1c", "1.1c", hl1_1c, hl2_1c, text_1c, graphics_1c),
        ]
    ),
    (
        ad_set_2_label, meta_desc_2,
        [
            ("2a", "2.1a", hl1_2a, hl2_2a, text_2a, graphics_2a),
            ("2b", "2.1b", hl1_2b, hl2_2b, text_2b, graphics_2b),
            ("2c", "2.1c", hl1_2c, hl2_2c, text_2c, graphics_2c),
        ]
    ),
    (
        ad_set_3_label, meta_desc_3,
        [
            ("3a", "3.1a", hl1_3a, hl2_3a, text_3a, graphics_3a),
            ("3b", "3.1b", hl1_3b, hl2_3b, text_3b, graphics_3b),
            ("3c", "3.1c", hl1_3c, hl2_3c, text_3c, graphics_3c),
        ]
    ),
    (
        ad_set_4_label, meta_desc_4,
        [
            ("4a", "4.1a", hl1_4a, hl2_4a, text_4a, graphics_4a),
            ("4b", "4.1b", hl1_4b, hl2_4b, text_4b, graphics_4b),
            ("4c", "4.1c", hl1_4c, hl2_4c, text_4c, graphics_4c),
        ]
    ),
    (
        ad_set_5_label, meta_desc_5,
        [
            ("5a", "5.1a", hl1_5a, hl2_5a, text_5a, graphics_5a),
            ("5b", "5.1b", hl1_5b, hl2_5b, text_5b, graphics_5b),
            ("5c", "5.1c", hl1_5c, hl2_5c, text_5c, graphics_5c),
        ]
    ),
]

current_row = 3  # start after header rows

for (adset_label, meta_desc, variations) in angles:

    # --- Spacer row before each AD SET ---
    ws.row_dimensions[current_row].height = 6
    current_row += 1

    # --- AD SET header row ---
    r = current_row
    set_cell(ws, r, 2, adset_label, font=font_bold, fill=fill_adset, alignment=align_top_wrap)
    set_cell(ws, r, 3, "ROUND 1",   font=font_bold, fill=fill_adset, alignment=align_top_wrap)
    set_cell(ws, r, 4, None,        fill=fill_adset, alignment=align_top_wrap)
    set_cell(ws, r, 5, "ROUND 2",   font=font_bold, fill=fill_adset, alignment=align_top_wrap)
    set_cell(ws, r, 6, None,        fill=fill_adset, alignment=align_top_wrap)
    set_cell(ws, r, 7, "ROUND 3",   font=font_bold, fill=fill_adset, alignment=align_top_wrap)
    set_cell(ws, r, 8, None,        fill=fill_adset, alignment=align_top_wrap)
    set_cell(ws, r, 9, "ROUND 1",   font=font_bold, fill=fill_adset, alignment=align_top_wrap)
    ws.row_dimensions[r].height = 20
    current_row += 1

    # --- Meta Description row ---
    r = current_row
    set_cell(ws, r, 2, "Meta Description", font=font_italic, fill=fill_metadesc, alignment=align_top_wrap)
    set_cell(ws, r, 3, meta_desc, font=font_italic, fill=fill_metadesc, alignment=align_top_wrap)
    set_cell(ws, r, 4, None, fill=fill_metadesc, alignment=align_top_wrap)
    set_cell(ws, r, 5, None, fill=fill_metadesc, alignment=align_top_wrap)
    set_cell(ws, r, 6, None, fill=fill_metadesc, alignment=align_top_wrap)
    set_cell(ws, r, 7, None, fill=fill_metadesc, alignment=align_top_wrap)
    set_cell(ws, r, 8, None, fill=fill_metadesc, alignment=align_top_wrap)
    set_cell(ws, r, 9, None, fill=fill_metadesc, alignment=align_top_wrap)
    ws.row_dimensions[r].height = 50
    current_row += 1

    # --- Variations ---
    for (var_id, round2_id, hl1, hl2, body_text, graphics_text) in variations:

        # Variation label row
        r = current_row
        set_cell(ws, r, 2, None,       font=font_bold, fill=fill_varlabel, alignment=align_top_wrap)
        set_cell(ws, r, 3, var_id,     font=font_bold, fill=fill_varlabel, alignment=align_top_wrap)
        set_cell(ws, r, 4, None,       fill=fill_varlabel, alignment=align_top_wrap)
        set_cell(ws, r, 5, round2_id,  font=font_bold, fill=fill_varlabel, alignment=align_top_wrap)
        set_cell(ws, r, 6, None,       fill=fill_varlabel, alignment=align_top_wrap)
        set_cell(ws, r, 7, None,       font=font_bold, fill=fill_varlabel, alignment=align_top_wrap)
        set_cell(ws, r, 8, None,       fill=fill_varlabel, alignment=align_top_wrap)
        set_cell(ws, r, 9, var_id,     font=font_bold, fill=fill_varlabel, alignment=align_top_wrap)
        ws.row_dimensions[r].height = 18
        current_row += 1

        # Meta Headline 1
        r = current_row
        set_cell(ws, r, 2, "Meta Headline 1", alignment=align_top_wrap)
        set_cell(ws, r, 3, hl1, alignment=align_top_wrap)
        ws.row_dimensions[r].height = 20
        current_row += 1

        # Meta Headline 2
        r = current_row
        set_cell(ws, r, 2, "Meta Headline 2", alignment=align_top_wrap)
        set_cell(ws, r, 3, hl2, alignment=align_top_wrap)
        ws.row_dimensions[r].height = 20
        current_row += 1

        # Meta Text
        r = current_row
        set_cell(ws, r, 2, "Meta Text", alignment=align_top_wrap)
        set_cell(ws, r, 3, body_text, alignment=align_top_wrap)
        set_cell(ws, r, 9, graphics_text, alignment=align_top_wrap)
        ws.row_dimensions[r].height = 200
        current_row += 1

        # Small spacer between variations
        ws.row_dimensions[current_row].height = 4
        current_row += 1


# ============================================================
# SAVE
# ============================================================
output_path = (
    "/Users/foxsss/Library/CloudStorage/GoogleDrive-miliescu92@gmail.com"
    "/My Drive/1_DBH/0_Clients/YOUnity - Impeccable Media"
    "/Younity Copywriting Agent/Blueprint/brg-cc-k3"
    "/brg-cc-k3 Facebook Ad Copy.xlsx"
)

wb.save(output_path)
print(f"Saved: {output_path}")
